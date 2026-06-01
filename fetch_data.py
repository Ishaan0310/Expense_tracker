"""
fetch_data.py  —  Finance Tracker (Project 1)
----------------------------------------------
Downloads the personal finance dataset from Kaggle, groups spending
by Category + Month to get Actual figures, auto-sets Budget as 10%
above average monthly spend per category, and writes budget_input.xlsx
so tracker.py can run on real data.

Dataset: https://www.kaggle.com/datasets/bukolafatunde/personal-finance
         ~1,000 rows of real personal expense records

Setup needed before running:
  1. pip install -r requirements.txt
  2. Kaggle API key at ~/.kaggle/kaggle.json  (see README)

Run: python3 fetch_data.py
"""

import os
import sys
import glob

import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import kaggle


# ── CONFIG ────────────────────────────────────────────────────────────────────
# To swap datasets: change DATASET_ID and the four column-name variables below.
# Everything else adjusts automatically.

DATASET_ID    = "bukolafatunde/personal-finance"   # Kaggle dataset identifier
DOWNLOAD_DIR  = "./raw_data"                        # where the zip lands
OUTPUT_FILE   = "budget_input.xlsx"                 # what tracker.py reads

# Column names in the downloaded CSV
# → If you switch datasets, check its column headers and update these four lines
COL_DATE      = "Date"
COL_CATEGORY  = "Category"
COL_AMOUNT    = "Amount"
COL_TYPE      = "Type"           # set to None if file has only expenses

# The value in COL_TYPE that means "this row is an expense"
EXPENSE_VALUE = "Expense"

# Budget = average monthly spend per category × this factor
# 1.10 → budget is 10% higher than average → a realistic but firm target
BUDGET_FACTOR = 1.10
# ─────────────────────────────────────────────────────────────────────────────


def download():
    os.makedirs(DOWNLOAD_DIR, exist_ok=True)
    print(f"downloading  →  {DATASET_ID}")
    kaggle.api.authenticate()
    kaggle.api.dataset_download_files(
        DATASET_ID, path=DOWNLOAD_DIR, unzip=True, quiet=False
    )
    print("download done.\n")


def find_csv():
    # sometimes a dataset unzips to multiple CSVs — pick the largest one
    csvs = glob.glob(os.path.join(DOWNLOAD_DIR, "**/*.csv"), recursive=True)
    if not csvs:
        print("no CSV files found after download.")
        print(f"check {DOWNLOAD_DIR}/ manually and confirm the dataset extracted correctly.")
        sys.exit(1)
    picked = max(csvs, key=os.path.getsize)
    print(f"using file: {picked}")
    return picked


def load_and_clean(path):
    df = pd.read_csv(path, parse_dates=[COL_DATE])

    # keep only expenses if a type column exists
    if COL_TYPE and COL_TYPE in df.columns:
        df = df[df[COL_TYPE].str.strip().str.lower() == EXPENSE_VALUE.lower()]

    # drop unusable rows
    df = df.dropna(subset=[COL_AMOUNT, COL_CATEGORY])
    df[COL_AMOUNT] = pd.to_numeric(df[COL_AMOUNT], errors="coerce").abs()
    df = df[df[COL_AMOUNT] > 0]

    # month name for grouping (January, February …)
    df["Month"] = df[COL_DATE].dt.strftime("%B")

    print(f"{len(df)} expense rows  |  {df['Month'].nunique()} months  |  {df[COL_CATEGORY].nunique()} categories")
    return df


def build_budget_vs_actual(df):
    # actual = total spend per category per month
    actuals = (
        df.groupby([COL_CATEGORY, "Month"], as_index=False)[COL_AMOUNT]
        .sum()
        .rename(columns={COL_CATEGORY: "Category", COL_AMOUNT: "Actual"})
    )
    actuals["Actual"] = actuals["Actual"].round(2)

    # budget = (average of all months for that category) × budget factor
    avg_per_cat = actuals.groupby("Category")["Actual"].mean()
    actuals["Budgeted"] = (actuals["Category"].map(avg_per_cat) * BUDGET_FACTOR).round(2)

    # drop months that have very few categories (likely partial data)
    month_counts = actuals.groupby("Month")["Category"].count()
    good_months  = month_counts[month_counts >= 3].index
    actuals       = actuals[actuals["Month"].isin(good_months)]

    actuals = actuals[["Category", "Month", "Budgeted", "Actual"]]
    print(f"output: {len(actuals)} rows across {actuals['Month'].nunique()} months")
    return actuals


def write_excel(df):
    wb  = openpyxl.Workbook()
    ws  = wb.active
    ws.title = "Budget"

    hdr_fill = PatternFill("solid", fgColor="2B4590")
    hdr_font = Font(bold=True, color="FFFFFF", size=11)
    alt_fill = PatternFill("solid", fgColor="F5F7FF")
    bdr = Border(
        left=Side(style="thin", color="D0D0D0"),
        right=Side(style="thin", color="D0D0D0"),
        top=Side(style="thin", color="D0D0D0"),
        bottom=Side(style="thin", color="D0D0D0"),
    )

    for col, h in enumerate(["Category", "Month", "Budgeted", "Actual"], 1):
        c = ws.cell(row=1, column=col, value=h)
        c.fill = hdr_fill
        c.font = hdr_font
        c.alignment = Alignment(horizontal="center")
        c.border = bdr
    ws.row_dimensions[1].height = 22

    for ri, (_, row) in enumerate(df.iterrows(), start=2):
        for ci, val in enumerate([row["Category"], row["Month"], row["Budgeted"], row["Actual"]], 1):
            c = ws.cell(row=ri, column=ci, value=val)
            c.border = bdr
            c.alignment = Alignment(horizontal="left" if ci == 1 else "center")
            if ri % 2 == 0:
                c.fill = alt_fill
            if ci in [3, 4]:
                c.number_format = "#,##0.00"

    for col, width in zip("ABCD", [22, 13, 14, 14]):
        ws.column_dimensions[col].width = width

    wb.save(OUTPUT_FILE)
    print(f"\nfile saved  →  {OUTPUT_FILE}")
    print("tip: open it and manually adjust the Budgeted column to your own targets if needed.")
    print("     then run:  python3 tracker.py")


def main():
    download()
    csv_path = find_csv()
    df       = load_and_clean(csv_path)
    out_df   = build_budget_vs_actual(df)
    write_excel(out_df)


if __name__ == "__main__":
    main()
