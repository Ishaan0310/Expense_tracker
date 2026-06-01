"""
Personal Finance Variance Tracker
----------------------------------
Reads budget vs actual data from budget_input.xlsx,
computes variance %, flags anything 40%+ over budget as an anomaly,
and writes a formatted report back to budget_report.xlsx.

Run: python3 tracker.py
"""

import sys
import datetime
import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference


INPUT_FILE = "budget_input.xlsx"
OUTPUT_FILE = "budget_report.xlsx"

# anything over this threshold gets flagged as an anomaly
OVERSPEND_THRESHOLD = 40  # percent


# ── helpers ────────────────────────────────────────────────────────────────────

def thin_border(color="CCCCCC"):
    s = Side(style="thin", color=color)
    return Border(left=s, right=s, top=s, bottom=s)


def apply_header_style(cell, bg_hex, font_color="FFFFFF"):
    cell.fill = PatternFill("solid", fgColor=bg_hex)
    cell.font = Font(bold=True, color=font_color, size=11)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = thin_border()


# ── data loading & calculation ──────────────────────────────────────────────────

def load_budget(filepath):
    try:
        df = pd.read_excel(filepath, sheet_name="Budget")
    except FileNotFoundError:
        print(f"error: can't find {filepath}")
        print("run create_sample_data.py first to generate a sample input file.")
        sys.exit(1)
    except Exception as e:
        print(f"error reading file: {e}")
        sys.exit(1)

    needed = {"Category", "Month", "Budgeted", "Actual"}
    missing = needed - set(df.columns)
    if missing:
        print(f"missing columns in Budget sheet: {missing}")
        sys.exit(1)

    # clean up
    df["Budgeted"] = pd.to_numeric(df["Budgeted"], errors="coerce").fillna(0)
    df["Actual"] = pd.to_numeric(df["Actual"], errors="coerce").fillna(0)
    df["Month"] = df["Month"].astype(str).str.strip()
    df["Category"] = df["Category"].astype(str).str.strip()

    return df


def compute_variance(df):
    df = df.copy()

    df["Variance"] = df["Actual"] - df["Budgeted"]

    # avoid divide by zero
    df["Variance_%"] = df.apply(
        lambda r: round((r["Actual"] - r["Budgeted"]) / r["Budgeted"] * 100, 2)
        if r["Budgeted"] != 0 else 0.0,
        axis=1
    )

    df["Flagged"] = df["Variance_%"] > OVERSPEND_THRESHOLD

    df["Status"] = df.apply(
        lambda r: "OVER BUDGET"
        if r["Flagged"]
        else ("Under budget" if r["Variance"] < 0 else "On track"),
        axis=1
    )

    return df


# ── sheet writers ───────────────────────────────────────────────────────────────

def write_summary(wb, df):
    sheet_name = "Summary"
    if sheet_name in wb.sheetnames:
        del wb[sheet_name]
    ws = wb.create_sheet(sheet_name, 0)

    red   = PatternFill("solid", fgColor="FFD6D6")
    green = PatternFill("solid", fgColor="D6F5D6")
    blue  = PatternFill("solid", fgColor="F0F4FF")

    # title row
    ws.merge_cells("A1:G1")
    ws["A1"].value = (
        f"Monthly Budget Variance Report  —  "
        f"{datetime.date.today().strftime('%B %Y')}"
    )
    ws["A1"].font = Font(bold=True, size=14, color="1A3A6B")
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 34
    ws.row_dimensions[2].height = 6

    # column headers
    cols = ["Category", "Month", "Budgeted (₹)", "Actual (₹)", "Variance (₹)", "Variance %", "Status"]
    for i, col in enumerate(cols, 1):
        cell = ws.cell(row=3, column=i, value=col)
        apply_header_style(cell, "2B4590")
    ws.row_dimensions[3].height = 22

    # data
    for row_i, (_, row) in enumerate(df.iterrows(), start=4):
        row_vals = [
            row["Category"], row["Month"],
            row["Budgeted"], row["Actual"],
            row["Variance"], row["Variance_%"],
            row["Status"]
        ]
        for col_i, val in enumerate(row_vals, 1):
            cell = ws.cell(row=row_i, column=col_i, value=val)
            cell.border = thin_border()
            cell.alignment = Alignment(
                horizontal="left" if col_i == 1 else "center",
                vertical="center"
            )

            # row color
            if row["Flagged"]:
                cell.fill = red
            elif row["Variance"] < 0:
                cell.fill = green
            elif row_i % 2 == 0:
                cell.fill = blue

            # number formats
            if col_i in [3, 4, 5]:
                cell.number_format = "#,##0.00"
            elif col_i == 6:
                cell.number_format = '0.00"%"'

    # column widths
    for i, w in enumerate([20, 13, 16, 16, 16, 13, 16], 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    return ws


def write_anomalies(wb, df):
    sheet_name = "Anomalies"
    if sheet_name in wb.sheetnames:
        del wb[sheet_name]
    ws = wb.create_sheet(sheet_name)

    flagged = df[df["Flagged"]].copy()

    # header
    ws.merge_cells("A1:G1")
    ws["A1"].value = (
        f"Overspend Anomalies  —  categories more than "
        f"{OVERSPEND_THRESHOLD}% over budget"
    )
    ws["A1"].font = Font(bold=True, size=13, color="A93226")
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30
    ws.row_dimensions[2].height = 6

    if flagged.empty:
        ws["A3"].value = "No anomalies this month — everything looks fine."
        ws["A3"].font = Font(italic=True, color="27AE60", size=12)
        return ws

    cols = ["Category", "Month", "Budgeted (₹)", "Actual (₹)", "Variance (₹)", "Variance %", "Status"]
    for i, col in enumerate(cols, 1):
        cell = ws.cell(row=3, column=i, value=col)
        apply_header_style(cell, "C0392B")
    ws.row_dimensions[3].height = 22

    red = PatternFill("solid", fgColor="FFD6D6")
    for row_i, (_, row) in enumerate(flagged.iterrows(), start=4):
        row_vals = [
            row["Category"], row["Month"],
            row["Budgeted"], row["Actual"],
            row["Variance"], row["Variance_%"], row["Status"]
        ]
        for col_i, val in enumerate(row_vals, 1):
            cell = ws.cell(row=row_i, column=col_i, value=val)
            cell.fill = red
            cell.border = thin_border()
            cell.alignment = Alignment(
                horizontal="left" if col_i == 1 else "center",
                vertical="center"
            )
            if col_i in [3, 4, 5]:
                cell.number_format = "#,##0.00"
            elif col_i == 6:
                cell.number_format = '0.00"%"'

    for i, w in enumerate([20, 13, 16, 16, 16, 13, 16], 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    return ws


def write_category_pivot(wb, df):
    sheet_name = "By_Category"
    if sheet_name in wb.sheetnames:
        del wb[sheet_name]
    ws = wb.create_sheet(sheet_name)

    pivot = (
        df.groupby("Category", as_index=False)
        .agg(Total_Budgeted=("Budgeted", "sum"), Total_Actual=("Actual", "sum"))
    )
    pivot["Net_Variance"] = pivot["Total_Actual"] - pivot["Total_Budgeted"]
    pivot["Variance_%"] = (
        (pivot["Total_Actual"] - pivot["Total_Budgeted"]) / pivot["Total_Budgeted"] * 100
    ).round(2)
    pivot = pivot.sort_values("Variance_%", ascending=False)

    red   = PatternFill("solid", fgColor="FFD6D6")
    green = PatternFill("solid", fgColor="D6F5D6")

    ws.merge_cells("A1:E1")
    ws["A1"].value = "Category-level Spending Summary (all months combined)"
    ws["A1"].font = Font(bold=True, size=13, color="1A5276")
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30
    ws.row_dimensions[2].height = 6

    cols = ["Category", "Total Budgeted (₹)", "Total Actual (₹)", "Net Variance (₹)", "Variance %"]
    for i, col in enumerate(cols, 1):
        cell = ws.cell(row=3, column=i, value=col)
        apply_header_style(cell, "1A5276")
    ws.row_dimensions[3].height = 22

    for row_i, (_, row) in enumerate(pivot.iterrows(), start=4):
        row_vals = [
            row["Category"], row["Total_Budgeted"],
            row["Total_Actual"], row["Net_Variance"], row["Variance_%"]
        ]
        for col_i, val in enumerate(row_vals, 1):
            cell = ws.cell(row=row_i, column=col_i, value=val)
            cell.border = thin_border()
            cell.alignment = Alignment(
                horizontal="left" if col_i == 1 else "center",
                vertical="center"
            )
            if col_i in [2, 3, 4]:
                cell.number_format = "#,##0.00"
            elif col_i == 5:
                cell.number_format = '0.00"%"'

            if row["Variance_%"] > OVERSPEND_THRESHOLD:
                cell.fill = red
            elif row["Variance_%"] < 0:
                cell.fill = green

    for i, w in enumerate([20, 20, 20, 20, 13], 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # bar chart — budget vs actual by category
    chart = BarChart()
    chart.type = "col"
    chart.title = "Budget vs Actual by Category"
    chart.y_axis.title = "Amount (₹)"
    chart.style = 10
    chart.width = 22
    chart.height = 14

    n = len(pivot)
    data_ref = Reference(ws, min_col=2, max_col=3, min_row=3, max_row=3 + n)
    cats_ref = Reference(ws, min_col=1, min_row=4, max_row=3 + n)
    chart.add_data(data_ref, titles_from_data=True)
    chart.set_categories(cats_ref)
    ws.add_chart(chart, f"A{6 + n}")

    return ws


# ── main ─────────────────────────────────────────────────────────────────────

def main():
    print(f"reading {INPUT_FILE}...")
    df = load_budget(INPUT_FILE)

    print("computing variances...")
    df = compute_variance(df)

    flagged_count = df["Flagged"].sum()
    print(f"found {flagged_count} anomaly/anomalies (>{OVERSPEND_THRESHOLD}% over budget)")

    # load the workbook to keep the original Budget sheet
    wb = openpyxl.load_workbook(INPUT_FILE)

    print("writing Summary sheet...")
    write_summary(wb, df)

    print("writing Anomalies sheet...")
    write_anomalies(wb, df)

    print("writing By_Category sheet...")
    write_category_pivot(wb, df)

    wb.save(OUTPUT_FILE)

    print(f"\nreport saved → {OUTPUT_FILE}")
    print(f"sheets: Summary | Anomalies | By_Category | Budget (original)\n")

    if flagged_count > 0:
        print(f"overspend details ({flagged_count} flagged):")
        flagged = df[df["Flagged"]][["Category", "Month", "Budgeted", "Actual", "Variance_%"]]
        for _, r in flagged.iterrows():
            sign = "+" if r["Variance_%"] > 0 else ""
            print(f"  {r['Category']:18s}  {r['Month']:10s}  {sign}{r['Variance_%']:.1f}% over budget")
    else:
        print("no anomalies found — all categories within threshold.")


if __name__ == "__main__":
    main()
